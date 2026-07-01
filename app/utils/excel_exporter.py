import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.testcase import TestCaseResult

_HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
_DATA_FILL_ODD  = PatternFill("solid", fgColor="EBF5FB")
_DATA_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
_BORDER_SIDE    = Side(style="thin", color="AAAAAA")
_BORDER         = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# Columns match the original VF34_update.xlsx "Body_Testcases" sheet.
# AI-specific metadata columns are appended at the end.
HEADERS = [
    "ID CB",                            #  1
    "Domain",                           #  2
    "FIP References",                   #  3
    "Functional Requirement Category",  #  4
    "Priority",                         #  5
    "Name",                             #  6
    "Description",                      #  7
    "Description.1",                    #  8
    "Pre-Action",                       #  9
    "Post-Action",                      # 10
    "Test Steps.Action",                # 11
    "Test Steps.Expected result",       # 12
    "Test Steps.Critical",              # 13
    "Test Steps.Output on CAN signal",  # 14
    "Status",                           # 15
    "Type",                             # 16
    "Environment",                      # 17
    "Vehicle Status",                   # 18
    "Power State",                      # 19
    "Creator",                          # 20
    "Test Purpose",                     # 21
    "Market",                           # 22
    "Variant",                          # 23
    "FIP No.",                          # 24
    "FIP Test item",                    # 25
    "Test design technical",            # 26
    "Note",                             # 27
    "Updater",                          # 28
    "Create date",                      # 29
    "Update date",                      # 30
    "version",                          # 31
    # --- AI metadata ---
    "Generation Path",                  # 32
    "Coverage Score",                   # 33
    "Classification",                   # 34
    "Best Match TC",                    # 35
    "Final Expected Result",            # 36
]

COL_WIDTHS = [
    20, 18, 30, 35, 10, 45, 55, 20, 35, 25,
    55, 55, 14, 30, 10, 14, 14, 14, 25, 20,
    45, 15, 15, 15, 30, 28, 20, 15, 14, 14, 8,
    16, 14, 18, 40, 55,
]


def _combine_steps(steps, field: str) -> str:
    """
    Combine TestStep list into a single numbered cell value matching the
    original format:
        1. <step 1 text>

        2. <step 2 text>

        3. <step 3 text>
    """
    if not steps:
        return ""
    lines = []
    for step in steps:
        text = getattr(step, field, "").strip()
        n = step.step_number
        # Prefix with number if not already present
        if not text.startswith(f"{n}."):
            text = f"{n}. {text}"
        lines.append(text)
    return "\n\n".join(lines)


def _write_header(ws) -> None:
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28


def _apply_row_style(ws, row: int, fill: PatternFill) -> None:
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.border    = _BORDER
        cell.alignment = _LEFT


def build_excel(results: list[TestCaseResult]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Body_Testcases"
    ws.freeze_panes = "A2"

    _write_header(ws)

    today = date.today().isoformat()
    current_row = 2

    for idx, item in enumerate(results):
        tc    = item.testcase
        cov   = item.coverage
        steps = tc.steps or []

        best_match_title = ""
        if cov.best_match:
            best_match_title = cov.best_match.payload.get("title", str(cov.best_match.qdrant_point_id))

        fill = _DATA_FILL_ODD if idx % 2 == 0 else _DATA_FILL_EVEN
        _apply_row_style(ws, current_row, fill)

        preconditions_text = "\n".join(tc.preconditions)

        # --- Original columns ---
        ws.cell(current_row,  1, tc.testcase_id)
        ws.cell(current_row,  2, tc.domain)
        ws.cell(current_row,  3, "")                                    # FIP References
        ws.cell(current_row,  4, "")                                    # Functional Req Category
        ws.cell(current_row,  5, "High")
        ws.cell(current_row,  6, tc.title)                              # Name
        ws.cell(current_row,  7, item.scenario_text)                    # Description
        ws.cell(current_row,  8, "")                                    # Description.1
        ws.cell(current_row,  9, preconditions_text)                    # Pre-Action
        ws.cell(current_row, 10, "")                                    # Post-Action
        ws.cell(current_row, 11, _combine_steps(steps, "action"))       # Test Steps.Action
        ws.cell(current_row, 12, _combine_steps(steps, "expected_result"))  # Test Steps.Expected result
        ws.cell(current_row, 13, "False")
        ws.cell(current_row, 14, "")                                    # CAN signal
        ws.cell(current_row, 15, "New")
        ws.cell(current_row, 16, "Functional")
        ws.cell(current_row, 17, "Vehicle")
        ws.cell(current_row, 18, "Static")
        ws.cell(current_row, 19, "")                                    # Power State
        ws.cell(current_row, 20, "AI Generated")
        ws.cell(current_row, 21, tc.title)                              # Test Purpose
        ws.cell(current_row, 22, "")                                    # Market
        ws.cell(current_row, 23, "")                                    # Variant
        ws.cell(current_row, 24, "")                                    # FIP No.
        ws.cell(current_row, 25, "")                                    # FIP Test item
        ws.cell(current_row, 26, "Equivalence Partitioning (EP)")
        ws.cell(current_row, 27, "")                                    # Note
        ws.cell(current_row, 28, "")                                    # Updater
        ws.cell(current_row, 29, today)
        ws.cell(current_row, 30, today)
        ws.cell(current_row, 31, "1.0")
        # --- AI metadata ---
        ws.cell(current_row, 32, tc.generation_path)
        ws.cell(current_row, 33, round(tc.coverage_score, 2))
        ws.cell(current_row, 34, tc.coverage_classification)
        ws.cell(current_row, 35, best_match_title)
        ws.cell(current_row, 36, tc.final_expected_result)

        # Row height scales with the number of steps so content is visible
        ws.row_dimensions[current_row].height = max(40, 22 * max(len(steps), 1))

        current_row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
